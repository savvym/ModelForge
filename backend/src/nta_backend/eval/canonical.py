from __future__ import annotations

import ast
import json
from dataclasses import dataclass
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field

MAX_EVAL_DATASET_ROWS = 1000


class EvalMessage(BaseModel):
    role: Literal["system", "user", "assistant", "tool"]
    content: str = Field(min_length=1)


class EvalSampleInput(BaseModel):
    system: str | None = None
    messages: list[EvalMessage] = Field(min_length=1)
    prompt: str | None = None
    parameters: dict[str, Any] | None = None


class EvalSampleReference(BaseModel):
    answer: str | None = None
    answers: list[str] | None = None
    label: str | None = None
    reference_text: str | None = None


class EvalSampleScoring(BaseModel):
    method: Literal["accuracy", "exact-match", "rule-based", "judge-model"] = "accuracy"
    weight: float = Field(default=1.0, gt=0)
    rubric_id: str | None = None
    rubric_overrides: dict[str, Any] | None = None


class EvalSample(BaseModel):
    sample_id: str = Field(min_length=1, max_length=255)
    task_type: Literal["single-turn", "multi-turn"] = "single-turn"
    input: EvalSampleInput
    reference: EvalSampleReference = Field(default_factory=EvalSampleReference)
    scoring: EvalSampleScoring = Field(default_factory=EvalSampleScoring)
    metadata: dict[str, Any] = Field(default_factory=dict)


@dataclass(frozen=True)
class NormalizedEvalDataset:
    format_name: str
    record_count: int
    samples: list[EvalSample]

    def to_jsonl_bytes(self) -> bytes:
        return dump_eval_samples_jsonl(self.samples)

    def schema_report(self) -> dict[str, Any]:
        return {
            "format": self.format_name,
            "record_count": self.record_count,
            "normalized": True,
            "warnings": [],
            "errors": [],
        }


def build_normalized_eval_artifact_key(source_object_key: str, artifact_name: str) -> str:
    source_path = PurePosixPath(source_object_key)
    return str(source_path.parent / "_normalized" / artifact_name)


def dump_eval_samples_jsonl(samples: list[EvalSample]) -> bytes:
    lines = [
        json.dumps(
            sample.model_dump(mode="json", exclude_none=True),
            ensure_ascii=False,
        )
        for sample in samples
    ]
    return ("\n".join(lines) + "\n").encode("utf-8")


def load_eval_samples_jsonl(body: bytes) -> list[EvalSample]:
    samples: list[EvalSample] = []
    for index, raw_line in enumerate(body.decode("utf-8", errors="replace").splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"第 {index} 行不是合法 JSON。") from exc
        samples.append(EvalSample.model_validate(payload))
    if not samples:
        raise ValueError("评测数据集不能为空。")
    validate_eval_samples(samples)
    return samples


def normalize_eval_dataset_bytes(file_name: str, body: bytes) -> NormalizedEvalDataset:
    lower_name = file_name.lower()
    if lower_name.endswith(".jsonl"):
        samples, format_name = _normalize_jsonl_dataset(body)
    elif lower_name.endswith(".xlsx"):
        samples, format_name = _normalize_spreadsheet_dataset(file_name, body, workbook_type="xlsx")
    elif lower_name.endswith(".xls"):
        samples, format_name = _normalize_spreadsheet_dataset(file_name, body, workbook_type="xls")
    else:
        raise ValueError("当前仅支持 JSONL、XLSX、XLS 评测数据集。")

    if not samples:
        raise ValueError("评测数据集不能为空。")

    if len(samples) > MAX_EVAL_DATASET_ROWS:
        raise ValueError(f"评测数据集单文件最多支持 {MAX_EVAL_DATASET_ROWS} 条样本。")

    validate_eval_samples(samples)

    return NormalizedEvalDataset(
        format_name=format_name,
        record_count=len(samples),
        samples=samples,
    )


def _normalize_jsonl_dataset(body: bytes) -> tuple[list[EvalSample], str]:
    normalized_samples: list[EvalSample] = []
    formats: set[str] = set()

    for line_number, payload in _iter_jsonl_objects(body):
        sample, format_name = _normalize_eval_record(payload, line_number)
        normalized_samples.append(sample)
        formats.add(format_name)

    format_name = formats.pop() if len(formats) == 1 else "mixed-jsonl"
    return normalized_samples, format_name


def _normalize_spreadsheet_dataset(
    file_name: str,
    body: bytes,
    *,
    workbook_type: Literal["xlsx", "xls"],
) -> tuple[list[EvalSample], str]:
    rows = (
        _iter_xlsx_rows(body)
        if workbook_type == "xlsx"
        else _iter_xls_rows(body)
    )
    if not rows:
        raise ValueError("评测数据集不能为空。")
    if len(rows) > MAX_EVAL_DATASET_ROWS:
        raise ValueError(f"评测数据集单文件最多支持 {MAX_EVAL_DATASET_ROWS} 行。")

    if _looks_like_multi_turn_sheet(rows):
        return _normalize_multi_turn_sheet(rows), f"{workbook_type}-multi-turn"
    return _normalize_single_turn_sheet(rows), f"{workbook_type}-single-turn"


def _iter_jsonl_objects(body: bytes) -> list[tuple[int, dict[str, Any]]]:
    objects: list[tuple[int, dict[str, Any]]] = []
    for index, raw_line in enumerate(body.decode("utf-8", errors="replace").splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        if len(objects) >= MAX_EVAL_DATASET_ROWS:
            raise ValueError(f"评测数据集单文件最多支持 {MAX_EVAL_DATASET_ROWS} 行。")
        try:
            payload = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"第 {index} 行不是合法 JSON。") from exc
        if not isinstance(payload, dict):
            raise ValueError(f"第 {index} 行必须是 JSON 对象。")
        objects.append((index, payload))
    return objects


def _iter_xlsx_rows(body: bytes) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
    worksheet = workbook.active
    return _sheet_rows_from_iterable(worksheet.iter_rows(values_only=True))


def _iter_xls_rows(body: bytes) -> list[tuple[int, dict[str, Any]]]:
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        raise ValueError("当前环境缺少 XLS 解析依赖，请安装 xlrd。") from exc

    workbook = xlrd.open_workbook(file_contents=body)
    worksheet = workbook.sheet_by_index(0)
    return _sheet_rows_from_iterable(
        [
            [
                worksheet.cell_value(row_index, column_index)
                for column_index in range(worksheet.ncols)
            ]
            for row_index in range(worksheet.nrows)
        ]
    )


def _sheet_rows_from_iterable(rows: Any) -> list[tuple[int, dict[str, Any]]]:
    header_row_index: int | None = None
    headers: list[str] = []
    normalized_rows: list[tuple[int, dict[str, Any]]] = []

    for row_index, raw_values in enumerate(rows, start=1):
        values = list(raw_values)
        if not any(_coerce_text(cell) for cell in values):
            continue
        if header_row_index is None:
            header_row_index = row_index
            headers = [
                _normalize_sheet_header(value, position)
                for position, value in enumerate(values, start=1)
            ]
            if not any(headers):
                raise ValueError("Excel 表头不能为空。")
            continue

        payload: dict[str, Any] = {}
        for header, cell in zip(headers, values, strict=False):
            if not header:
                continue
            payload[header] = _excel_cell_value(cell)
        if payload:
            normalized_rows.append((row_index, payload))

    return normalized_rows


def _normalize_sheet_header(value: Any, position: int) -> str:
    text = _coerce_text(value)
    if not text:
        return f"_column_{position}"
    lowered = text.strip().lower().replace(" ", "_")
    lowered = lowered.replace("-", "_")
    alias_map = {
        "system": "system_prompt",
        "systemprompt": "system_prompt",
        "system_prompt": "system_prompt",
        "query": "query",
        "prompt": "query",
        "reference_response": "reference_response",
        "answer": "reference_response",
        "response": "response",
        "session_id": "session_id",
        "parameters": "parameters",
        "messages": "messages",
        "message": "messages",
    }
    return alias_map.get(lowered, lowered)


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _looks_like_multi_turn_sheet(rows: list[tuple[int, dict[str, Any]]]) -> bool:
    seen: set[str] = set()
    duplicated = False
    for _, payload in rows:
        session_id = _coerce_text(payload.get("session_id"))
        if not session_id:
            continue
        if session_id in seen:
            duplicated = True
            break
        seen.add(session_id)
    return duplicated


def _normalize_single_turn_sheet(rows: list[tuple[int, dict[str, Any]]]) -> list[EvalSample]:
    samples: list[EvalSample] = []
    for row_number, payload in rows:
        sample = _normalize_volc_single_turn_record(payload, row_number)
        samples.append(sample)
    return samples


def _normalize_multi_turn_sheet(rows: list[tuple[int, dict[str, Any]]]) -> list[EvalSample]:
    sessions: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for row_number, payload in rows:
        session_id = _coerce_text(payload.get("session_id"))
        if not session_id:
            raise ValueError(f"第 {row_number} 行缺少 session_id，多轮 Excel 评测集必须填写。")
        sessions.setdefault(session_id, []).append((row_number, payload))

    samples: list[EvalSample] = []
    for session_id, session_rows in sessions.items():
        samples.append(_normalize_multi_turn_session_rows(session_id, session_rows))
    return samples


def _normalize_multi_turn_session_rows(
    session_id: str,
    session_rows: list[tuple[int, dict[str, Any]]],
) -> EvalSample:
    systems = {
        system_prompt
        for _, payload in session_rows
        for system_prompt in [_coerce_text(payload.get("system_prompt"))]
        if system_prompt
    }
    if len(systems) > 1:
        raise ValueError(f"session_id={session_id} 的 system_prompt 必须保持一致。")
    system_prompt = next(iter(systems), None)

    messages: list[EvalMessage] = []
    final_reference = EvalSampleReference()
    final_parameters: dict[str, Any] | None = None

    for index, (row_number, payload) in enumerate(session_rows):
        query = _coerce_text(payload.get("query"))
        if not query:
            raise ValueError(f"第 {row_number} 行缺少 query。")
        messages.append(EvalMessage(role="user", content=query))

        response = _coerce_text(payload.get("response"))
        if response:
            messages.append(EvalMessage(role="assistant", content=response))

        reference = _reference_from_payload(payload)
        if _reference_has_values(reference):
            if index != len(session_rows) - 1:
                raise ValueError(
                    f"session_id={session_id} 的参考答案只能出现在最后一行。"
                )
            final_reference = reference

        parameters = _coerce_parameters(payload.get("parameters"), row_number=row_number)
        if parameters:
            final_parameters = parameters

    if not messages:
        raise ValueError(f"session_id={session_id} 不包含有效消息。")
    if messages[-1].role != "user":
        raise ValueError(f"session_id={session_id} 的最后一条消息必须是用户输入。")

    metadata = {
        "session_id": session_id,
        "source_format": "excel-multi-turn",
        "row_count": len(session_rows),
    }
    return EvalSample(
        sample_id=session_id,
        task_type="multi-turn",
        input=EvalSampleInput(
            system=system_prompt,
            messages=messages,
            prompt=_last_user_content(messages),
            parameters=final_parameters,
        ),
        reference=final_reference,
        scoring=EvalSampleScoring(),
        metadata=metadata,
    )


def _normalize_eval_record(payload: dict[str, Any], line_number: int) -> tuple[EvalSample, str]:
    if _looks_like_canonical_sample(payload):
        sample = EvalSample.model_validate(payload)
        return sample, "canonical-jsonl"

    if _looks_like_volc_multi_turn_record(payload):
        return _normalize_volc_multi_turn_record(payload, line_number), "volc-multi-turn-jsonl"

    if _looks_like_volc_single_turn_record(payload):
        return _normalize_volc_single_turn_record(payload, line_number), "volc-single-turn-jsonl"

    if isinstance(payload.get("input"), str):
        return _normalize_simple_qa_record(payload, line_number), "simple-qa-jsonl"

    if isinstance(payload.get("prompt"), str) or isinstance(payload.get("system"), str):
        return _normalize_mmlu_style_record(payload, line_number), "mmlu-jsonl"

    raise ValueError(f"第 {line_number} 行不符合支持的评测样本格式。")


def _looks_like_canonical_sample(payload: dict[str, Any]) -> bool:
    input_payload = payload.get("input")
    return (
        isinstance(payload.get("sample_id"), str)
        and isinstance(input_payload, dict)
        and isinstance(input_payload.get("messages"), list)
    )


def _looks_like_volc_single_turn_record(payload: dict[str, Any]) -> bool:
    return any(
        isinstance(payload.get(key), str)
        for key in ("query", "system_prompt", "reference_response")
    )


def _looks_like_volc_multi_turn_record(payload: dict[str, Any]) -> bool:
    messages = payload.get("messages")
    if not isinstance(messages, list):
        messages = payload.get("message")
    return isinstance(messages, list)


def _normalize_simple_qa_record(payload: dict[str, Any], line_number: int) -> EvalSample:
    prompt = _coerce_text(payload.get("input"))
    if not prompt:
        raise ValueError(f"第 {line_number} 行缺少 input。")

    return EvalSample(
        sample_id=_sample_id_from_payload(payload, line_number),
        task_type="single-turn",
        input=EvalSampleInput(
            messages=[EvalMessage(role="user", content=prompt)],
            prompt=prompt,
            parameters=_coerce_parameters(payload.get("parameters"), row_number=line_number),
        ),
        reference=_reference_from_payload(payload),
        scoring=EvalSampleScoring(method=_coerce_method(payload.get("method"))),
        metadata=_extra_metadata(
            payload,
            {
                "id",
                "sample_id",
                "input",
                "answer",
                "answers",
                "label",
                "reference_text",
                "method",
                "parameters",
            },
        ),
    )


def _normalize_mmlu_style_record(payload: dict[str, Any], line_number: int) -> EvalSample:
    prompt = _coerce_text(payload.get("prompt"))
    if not prompt:
        raise ValueError(f"第 {line_number} 行缺少 prompt。")
    system = _coerce_text(payload.get("system"))
    metadata = _extra_metadata(
        payload,
        {
            "id",
            "sample_id",
            "system",
            "prompt",
            "answer",
            "answers",
            "label",
            "reference_text",
            "reference_response",
            "method",
            "parameters",
        },
    )
    if "subject" not in metadata:
        subject = _infer_subject_from_prompt(prompt)
        if subject:
            metadata["subject"] = subject

    return EvalSample(
        sample_id=_sample_id_from_payload(payload, line_number),
        task_type="single-turn",
        input=EvalSampleInput(
            system=system,
            messages=[EvalMessage(role="user", content=prompt)],
            prompt=prompt,
            parameters=_coerce_parameters(payload.get("parameters"), row_number=line_number),
        ),
        reference=_reference_from_payload(payload),
        scoring=EvalSampleScoring(method=_coerce_method(payload.get("method"))),
        metadata=metadata,
    )


def _normalize_volc_single_turn_record(payload: dict[str, Any], line_number: int) -> EvalSample:
    query = _coerce_text(payload.get("query")) or _coerce_text(payload.get("prompt"))
    if not query:
        raise ValueError(f"第 {line_number} 行缺少 query/prompt。")
    system = _coerce_text(payload.get("system_prompt")) or _coerce_text(payload.get("system"))
    metadata = _extra_metadata(
        payload,
        {
            "id",
            "sample_id",
            "session_id",
            "system_prompt",
            "system",
            "query",
            "prompt",
            "answer",
            "answers",
            "label",
            "reference_text",
            "reference_response",
            "response",
            "method",
            "parameters",
        },
    )
    if payload.get("session_id") is not None:
        metadata["session_id"] = payload.get("session_id")

    return EvalSample(
        sample_id=_sample_id_from_payload(payload, line_number),
        task_type="single-turn",
        input=EvalSampleInput(
            system=system,
            messages=[EvalMessage(role="user", content=query)],
            prompt=query,
            parameters=_coerce_parameters(payload.get("parameters"), row_number=line_number),
        ),
        reference=_reference_from_payload(payload),
        scoring=EvalSampleScoring(method=_coerce_method(payload.get("method"))),
        metadata=metadata,
    )


def _normalize_volc_multi_turn_record(payload: dict[str, Any], line_number: int) -> EvalSample:
    raw_messages = payload.get("messages")
    if not isinstance(raw_messages, list):
        raw_messages = payload.get("message")
    if not isinstance(raw_messages, list) or not raw_messages:
        raise ValueError(f"第 {line_number} 行缺少 messages/message。")

    system_messages: list[str] = []
    messages: list[EvalMessage] = []
    for message_index, item in enumerate(raw_messages, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"第 {line_number} 行的第 {message_index} 条消息必须是对象。")
        role = _coerce_text(item.get("role"))
        content = _coerce_text(item.get("content"))
        if role not in {"system", "user", "assistant", "tool"} or not content:
            raise ValueError(f"第 {line_number} 行的第 {message_index} 条消息不合法。")
        if role == "system":
            system_messages.append(content)
            continue
        messages.append(EvalMessage(role=role, content=content))

    if not messages:
        raise ValueError(f"第 {line_number} 行缺少可用于推理的对话消息。")
    if messages[-1].role != "user":
        raise ValueError(f"第 {line_number} 行的 messages 必须以 user 结尾。")

    metadata = _extra_metadata(
        payload,
        {
            "id",
            "sample_id",
            "session_id",
            "messages",
            "message",
            "answer",
            "answers",
            "label",
            "reference_text",
            "reference_response",
            "method",
            "parameters",
        },
    )
    if payload.get("session_id") is not None:
        metadata["session_id"] = payload.get("session_id")

    return EvalSample(
        sample_id=_sample_id_from_payload(payload, line_number),
        task_type="multi-turn",
        input=EvalSampleInput(
            system="\n".join(system_messages) if system_messages else None,
            messages=messages,
            prompt=_last_user_content(messages),
            parameters=_coerce_parameters(payload.get("parameters"), row_number=line_number),
        ),
        reference=_reference_from_payload(payload),
        scoring=EvalSampleScoring(method=_coerce_method(payload.get("method"))),
        metadata=metadata,
    )


def _sample_id_from_payload(payload: dict[str, Any], line_number: int) -> str:
    for key in ("sample_id", "id", "session_id"):
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
        if isinstance(value, int):
            return str(value)
    return f"sample-{line_number:06d}"


def _reference_from_payload(payload: dict[str, Any]) -> EvalSampleReference:
    return EvalSampleReference(
        answer=(
            _coerce_text(payload.get("answer"))
            or _coerce_text(payload.get("reference_response"))
        ),
        answers=_coerce_text_list(payload.get("answers")),
        label=_coerce_text(payload.get("label")),
        reference_text=_coerce_text(payload.get("reference_text")),
    )


def _reference_has_values(reference: EvalSampleReference) -> bool:
    return bool(
        reference.answer
        or reference.answers
        or reference.label
        or reference.reference_text
    )


def _coerce_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip() or None


def _coerce_text_list(value: Any) -> list[str] | None:
    if value is None:
        return None
    if isinstance(value, list):
        items = [str(item).strip() for item in value if str(item).strip()]
        return items or None
    text = _coerce_text(value)
    if not text:
        return None
    return [text]


def _coerce_parameters(value: Any, *, row_number: int) -> dict[str, Any] | None:
    if value is None or value == "":
        return None
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            try:
                parsed = ast.literal_eval(text)
            except (SyntaxError, ValueError) as exc:
                raise ValueError(f"第 {row_number} 行的 parameters 不是合法字典。") from exc
        if not isinstance(parsed, dict):
            raise ValueError(f"第 {row_number} 行的 parameters 必须是对象。")
        return parsed
    raise ValueError(f"第 {row_number} 行的 parameters 必须是对象。")


def _coerce_method(value: Any) -> Literal["accuracy", "exact-match", "rule-based", "judge-model"]:
    if value == "rule-based":
        return "rule-based"
    if value == "judge-model":
        return "judge-model"
    if value == "exact-match":
        return "exact-match"
    return "accuracy"


def _extra_metadata(payload: dict[str, Any], consumed_keys: set[str]) -> dict[str, Any]:
    return {key: value for key, value in payload.items() if key not in consumed_keys}


def _infer_subject_from_prompt(prompt: str) -> str | None:
    for line in prompt.splitlines():
        if ":" not in line:
            continue
        head, tail = line.split(":", 1)
        if head.strip().lower() == "subject" and tail.strip():
            return tail.strip()
    return None


def _last_user_content(messages: list[EvalMessage]) -> str | None:
    for message in reversed(messages):
        if message.role == "user":
            return message.content
    return None


def validate_eval_samples(samples: list[EvalSample]) -> None:
    seen_sample_ids: set[str] = set()

    for index, sample in enumerate(samples, start=1):
        if sample.sample_id in seen_sample_ids:
            raise ValueError(f"样本 sample_id 重复：{sample.sample_id}")
        seen_sample_ids.add(sample.sample_id)

        requires_reference = _method_requires_reference(sample.scoring.method)
        has_reference = _reference_has_values(sample.reference)
        if requires_reference and not has_reference:
            raise ValueError(
                f"第 {index} 条样本缺少参考答案："
                f"{sample.sample_id} 使用 {sample.scoring.method} 时必须提供 reference。"
            )


def _method_requires_reference(method: str) -> bool:
    return method in {"accuracy", "exact-match", "rule-based"}
